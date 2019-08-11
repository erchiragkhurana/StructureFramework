import openpyxl

def dataGenerator():
    wk = openpyxl.load_workbook("E:/TestData.xlsx")
    sh = wk['Sheet1']
    r = sh.max_row
    li = []
    li1 = []
    for i in range(1,r+1):
        li1 = []
        un = sh.cell(i,1) #for username
        up = sh.cell(i,2) #for password
        li1.insert(0,un.value)
        li1.insert(1,up.value)
        li.insert(i-1,li1)

    print(li)
    return li